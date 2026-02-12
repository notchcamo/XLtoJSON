import json
import os
import re

import pandas as pd
from openpyxl import Workbook, load_workbook


# ---------------------------------------------------------------------------
# Excel header detection helpers
# ---------------------------------------------------------------------------

def _detect_xl_columns(ws):
    """Detect column types by inspecting headers and merged cells.

    Returns a dict with:
        simple        – list of (col_name, col_index) for plain columns
        flat_groups   – {base_name: [(col_name, col_index), ...]}
        struct_groups – {base_name: [{index, sub_fields, col_indices}, ...]}
        single_structs – set of base_names that have no _N suffix (single object)
        has_struct    – bool, True if any structured columns exist
    """
    # Build maps for row-1 merges
    h_merged = {}
    for rng in ws.merged_cells.ranges:
        if rng.min_row == 1 and rng.min_col != rng.max_col:
            for c in range(rng.min_col, rng.max_col + 1):
                h_merged[c] = (rng.min_col, rng.max_col)

    simple = []
    flat_groups: dict[str, list[tuple[str, int]]] = {}
    struct_groups: dict[str, list[dict]] = {}
    single_structs: set[str] = set()

    col = 1
    while col <= ws.max_column:
        header = ws.cell(1, col).value
        if header is None:
            col += 1
            continue

        header = str(header)
        m = re.match(r"^(.+)_(\d+)$", header)

        if col in h_merged:
            merge_start, merge_end = h_merged[col]
            sub_fields = []
            col_indices = []
            for sc in range(merge_start, merge_end + 1):
                sf = ws.cell(2, sc).value
                sub_fields.append(str(sf) if sf is not None else f"field_{sc}")
                col_indices.append(sc)

            if m:
                base = m.group(1)
                idx = int(m.group(2))
            else:
                base = header
                idx = 0
                single_structs.add(base)

            struct_groups.setdefault(base, []).append({
                "index": idx,
                "sub_fields": sub_fields,
                "col_indices": col_indices,
            })
            col = merge_end + 1
        elif m:
            base = m.group(1)
            flat_groups.setdefault(base, []).append((header, col))
            col += 1
        else:
            simple.append((header, col))
            col += 1

    for base in flat_groups:
        flat_groups[base].sort(key=lambda t: int(t[0][len(base) + 1:]))
    for base in struct_groups:
        struct_groups[base].sort(key=lambda d: d["index"])

    return {
        "simple": simple,
        "flat_groups": flat_groups,
        "struct_groups": struct_groups,
        "single_structs": single_structs,
        "has_struct": bool(struct_groups),
    }


# ---------------------------------------------------------------------------
# JSON column analysis helper
# ---------------------------------------------------------------------------

def _analyze_json_columns(df):
    """Classify DataFrame columns into simple / flat-array / structured-array /
    single-struct.

    Returns a dict with:
        simple        – [col_name, ...]
        flat_arrays   – {col_name: max_length}
        struct_arrays – {col_name: {"sub_fields": [...], "max_length": int}}
        single_structs – {col_name: {"sub_fields": [...]}}
    """
    simple = []
    flat_arrays: dict[str, int] = {}
    struct_arrays: dict[str, dict] = {}
    single_structs: dict[str, dict] = {}

    for col in df.columns:
        # Check for single dict (not in a list)
        has_dict = df[col].apply(lambda x: isinstance(x, dict)).any()
        if has_dict:
            all_keys: list[str] = []
            seen: set[str] = set()
            for val in df[col]:
                if isinstance(val, dict):
                    for k in val:
                        if k not in seen:
                            all_keys.append(k)
                            seen.add(k)
            single_structs[col] = {"sub_fields": all_keys}
            continue

        has_list = df[col].apply(lambda x: isinstance(x, list)).any()
        if not has_list:
            simple.append(col)
            continue

        # Check first non-empty list element
        is_struct = False
        for val in df[col]:
            if isinstance(val, list) and len(val) > 0:
                if isinstance(val[0], dict):
                    is_struct = True
                break

        expanded = df[col].apply(lambda x: x if isinstance(x, list) else [])
        max_len = int(expanded.apply(len).max())

        if is_struct:
            all_keys2: list[str] = []
            seen2: set[str] = set()
            for lst in expanded:
                for obj in lst:
                    if isinstance(obj, dict):
                        for k in obj:
                            if k not in seen2:
                                all_keys2.append(k)
                                seen2.add(k)
            struct_arrays[col] = {"sub_fields": all_keys2, "max_length": max_len}
        else:
            flat_arrays[col] = max_len

    return {
        "simple": simple,
        "flat_arrays": flat_arrays,
        "struct_arrays": struct_arrays,
        "single_structs": single_structs,
    }


# ---------------------------------------------------------------------------
# Excel → JSON
# ---------------------------------------------------------------------------

def convert_xl_to_json(xlsx_path: str, output_dir: str) -> None:
    """Convert Excel file to JSON, supporting simple, flat-array,
    structured-array, and single-object (merged-header) columns."""
    wb = load_workbook(xlsx_path)
    ws = wb.active
    info = _detect_xl_columns(ws)

    data_start = 3 if info["has_struct"] else 2

    records: list[dict] = []
    for row in range(data_start, ws.max_row + 1):
        rec: dict = {}

        for name, ci in info["simple"]:
            rec[name] = ws.cell(row, ci).value

        for base, cols in info["flat_groups"].items():
            vals = []
            for _, ci in cols:
                v = ws.cell(row, ci).value
                if v is not None:
                    vals.append(v)
            rec[base] = vals

        for base, elements in info["struct_groups"].items():
            arr: list[dict] = []
            for elem in elements:
                obj = {}
                for sf, ci in zip(elem["sub_fields"], elem["col_indices"]):
                    obj[sf] = ws.cell(row, ci).value
                if any(v is not None for v in obj.values()):
                    arr.append(obj)

            if base in info["single_structs"]:
                # Single object (no _N suffix) → unwrap from array
                rec[base] = arr[0] if arr else {}
            else:
                rec[base] = arr

        records.append(rec)

    os.makedirs(output_dir, exist_ok=True)
    base_name = os.path.splitext(os.path.basename(xlsx_path))[0]
    output_path = os.path.join(output_dir, f"{base_name}.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)


# ---------------------------------------------------------------------------
# JSON → Excel
# ---------------------------------------------------------------------------

def convert_json_to_xl(json_path: str, output_dir: str) -> None:
    """Convert JSON file to Excel, supporting simple, flat-array,
    structured-array, and single-object (merged-header) columns."""
    df = pd.read_json(json_path)
    analysis = _analyze_json_columns(df)
    has_struct = bool(analysis["struct_arrays"]) or bool(analysis["single_structs"])

    os.makedirs(output_dir, exist_ok=True)
    base_name = os.path.splitext(os.path.basename(json_path))[0]
    output_path = os.path.join(output_dir, f"{base_name}.xlsx")

    # Merge with existing file if present — use simple (scalar) columns as
    # the dedup key so that changed rows are *updated*, not duplicated.
    if os.path.exists(output_path):
        existing = _read_xl_as_records(output_path)
        new = json.loads(df.to_json(orient="records"))
        # Identify simple (non-list, non-dict) columns to use as merge key
        simple_keys = [
            k for k in (new[0] if new else {})
            if not isinstance((new[0] if new else {}).get(k), (list, dict))
        ]
        # Index existing records by simple-column values, then overlay new
        seen: dict[str, dict] = {}
        for rec in existing + new:
            key = json.dumps({k: rec.get(k) for k in simple_keys},
                             sort_keys=True, ensure_ascii=False)
            seen[key] = rec
        records = list(seen.values())
    else:
        records = json.loads(df.to_json(orient="records"))

    # Re-analyse from merged records
    df = pd.DataFrame(records)
    analysis = _analyze_json_columns(df)
    has_struct = bool(analysis["struct_arrays"]) or bool(analysis["single_structs"])

    wb = Workbook()
    ws = wb.active
    data_start = 3 if has_struct else 2

    # --- Write headers ---
    current_col = 1
    layout: list[tuple[str, str, object]] = []

    # Simple columns
    for col_name in analysis["simple"]:
        ws.cell(1, current_col, col_name)
        if has_struct:
            # Merge row 1-2 vertically so the header spans both header rows
            ws.merge_cells(
                start_row=1, start_column=current_col,
                end_row=2, end_column=current_col,
            )
        layout.append((col_name, "simple", current_col))
        current_col += 1

    # Flat array columns
    for col_name, max_len in analysis["flat_arrays"].items():
        start = current_col
        for i in range(max_len):
            ws.cell(1, current_col, f"{col_name}_{i}")
            if has_struct:
                ws.merge_cells(
                    start_row=1, start_column=current_col,
                    end_row=2, end_column=current_col,
                )
            current_col += 1
        layout.append((col_name, "flat", (start, max_len)))

    # Single-object structured columns (no _N suffix)
    for col_name, info in analysis["single_structs"].items():
        sub_fields = info["sub_fields"]
        merge_start = current_col
        merge_end = current_col + len(sub_fields) - 1
        ws.cell(1, merge_start, col_name)
        if merge_end > merge_start:
            ws.merge_cells(
                start_row=1, start_column=merge_start,
                end_row=1, end_column=merge_end,
            )
        for si, sf in enumerate(sub_fields):
            ws.cell(2, merge_start + si, sf)
        layout.append((col_name, "single_struct", (merge_start, sub_fields)))
        current_col = merge_end + 1

    # Array-of-objects structured columns (with _N suffix)
    for col_name, info in analysis["struct_arrays"].items():
        sub_fields = info["sub_fields"]
        max_len = info["max_length"]
        starts: list[tuple[int, list[str]]] = []
        for i in range(max_len):
            merge_start = current_col
            merge_end = current_col + len(sub_fields) - 1
            ws.cell(1, merge_start, f"{col_name}_{i}")
            if merge_end > merge_start:
                ws.merge_cells(
                    start_row=1, start_column=merge_start,
                    end_row=1, end_column=merge_end,
                )
            for si, sf in enumerate(sub_fields):
                ws.cell(2, merge_start + si, sf)
            starts.append((merge_start, sub_fields))
            current_col = merge_end + 1
        layout.append((col_name, "struct", starts))

    # --- Write data rows ---
    for ri, rec in enumerate(records):
        excel_row = data_start + ri
        for col_name, col_type, meta in layout:
            val = rec.get(col_name)
            if col_type == "simple":
                ws.cell(excel_row, meta, val)
            elif col_type == "flat":
                start, max_len = meta
                items = val if isinstance(val, list) else []
                for i in range(max_len):
                    ws.cell(excel_row, start + i, items[i] if i < len(items) else None)
            elif col_type == "single_struct":
                start, sub_fields = meta
                obj = val if isinstance(val, dict) else {}
                for si, sf in enumerate(sub_fields):
                    ws.cell(excel_row, start + si, obj.get(sf))
            elif col_type == "struct":
                items = val if isinstance(val, list) else []
                for obj_i, (start, sub_fields) in enumerate(meta):
                    obj = items[obj_i] if obj_i < len(items) and isinstance(items[obj_i], dict) else {}
                    for si, sf in enumerate(sub_fields):
                        ws.cell(excel_row, start + si, obj.get(sf))

    wb.save(output_path)


def _read_xl_as_records(xlsx_path: str) -> list[dict]:
    """Read an Excel file (with possible merged headers) back into records."""
    wb = load_workbook(xlsx_path)
    ws = wb.active
    info = _detect_xl_columns(ws)
    data_start = 3 if info["has_struct"] else 2

    records: list[dict] = []
    for row in range(data_start, ws.max_row + 1):
        rec: dict = {}
        for name, ci in info["simple"]:
            rec[name] = ws.cell(row, ci).value
        for base, cols in info["flat_groups"].items():
            vals = []
            for _, ci in cols:
                v = ws.cell(row, ci).value
                if v is not None:
                    vals.append(v)
            rec[base] = vals
        for base, elements in info["struct_groups"].items():
            arr: list[dict] = []
            for elem in elements:
                obj = {}
                for sf, ci in zip(elem["sub_fields"], elem["col_indices"]):
                    obj[sf] = ws.cell(row, ci).value
                if any(v is not None for v in obj.values()):
                    arr.append(obj)
            if base in info["single_structs"]:
                rec[base] = arr[0] if arr else {}
            else:
                rec[base] = arr
        records.append(rec)
    return records
